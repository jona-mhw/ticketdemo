from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, make_response, send_file
from flask_login import login_required, current_user
from models import db, Ticket, Patient, Surgery, Technique, StayAdjustmentCriterion, FpaModification, StandardizedReason, Doctor, DischargeTimeSlot
from datetime import datetime, timedelta
import json
import io
import csv
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .utils import _build_tickets_query, calculate_time_remaining, apply_sorting_to_query

tickets_bp = Blueprint('tickets', __name__)

def generate_ticket_id():
    """Generate unique ticket ID in format TH-YYYY-XXX"""
    current_year = datetime.now().year
    
    year_prefix = f"TH-{current_year}-"
    last_ticket = Ticket.query.filter(
        Ticket.id.like(f"{year_prefix}%"),
        Ticket.clinic_id == current_user.clinic_id
    ).order_by(Ticket.id.desc()).first()
    
    if last_ticket:
        last_number = int(last_ticket.id.split('-')[-1])
        new_number = last_number + 1
    else:
        new_number = 1
    
    return f"TH-{current_year}-{new_number:03d}"

def calculate_time_remaining(fpa):
    """Calculate detailed time remaining until FPA."""
    if not fpa:
        return None
    
    now = datetime.now()
    if fpa <= now:
        return {'days': 0, 'hours': 0, 'minutes': 0, 'seconds': 0, 'expired': True}
    
    time_diff = fpa - now
    days = time_diff.days
    hours, remainder = divmod(time_diff.seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    
    return {
        'days': days,
        'hours': hours,
        'minutes': minutes,
        'seconds': seconds,
        'expired': False
    }

@tickets_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create():
    clinic_id = current_user.clinic_id
    if request.method == 'POST':
        try:
            # Patient data
            rut = request.form.get('rut', '').strip()
            primer_nombre = request.form.get('primer_nombre', '').strip()
            apellido_paterno = request.form.get('apellido_paterno', '').strip()
            age = int(request.form.get('age', 0))
            sex = request.form.get('sex', '')
            episode_id = request.form.get('episode_id', '').strip()
            room_location = request.form.get('room_location', '').strip()
            
            # Surgery data
            surgery_id = int(request.form.get('surgery_id'))
            technique_id = int(request.form.get('technique_id'))
            pavilion_end_time = datetime.strptime(request.form.get('pavilion_end_time'), '%Y-%m-%dT%H:%M')
            
            doctor_id = request.form.get('doctor_id')
            doctor_id = int(doctor_id) if doctor_id else None
            discharge_slot_id = request.form.get('discharge_slot_id')
            discharge_slot_id = int(discharge_slot_id) if discharge_slot_id else None
            
            adjustment_ids = [int(id) for id in request.form.getlist('adjustment_ids') if id]
            
            if not all([rut, primer_nombre, apellido_paterno, age, sex, surgery_id, technique_id]):
                flash('Todos los campos obligatorios deben ser completados.', 'error')
                return redirect(url_for('tickets.create'))
            
            if episode_id:
                existing_ticket = Ticket.query.join(Patient).filter(
                    Patient.episode_id == episode_id,
                    Ticket.status == 'Vigente',
                    Ticket.clinic_id == clinic_id
                ).first()
                if existing_ticket:
                    flash(f'Ya existe un ticket vigente para el episodio {episode_id}: {existing_ticket.id}', 'error')
                    return redirect(url_for('tickets.create'))
            
            technique = Technique.query.filter_by(id=technique_id, clinic_id=clinic_id).first_or_404()
            surgery = Surgery.query.filter_by(id=surgery_id, clinic_id=clinic_id).first_or_404()
            
            adjustment_hours = 0
            if adjustment_ids:
                adjustments = StayAdjustmentCriterion.query.filter(StayAdjustmentCriterion.id.in_(adjustment_ids), StayAdjustmentCriterion.clinic_id == clinic_id).all()
                adjustment_hours = sum(adj.hours_adjustment for adj in adjustments)
            
            patient = Patient.query.filter_by(rut=rut, clinic_id=clinic_id).first()
            if not patient:
                patient = Patient(
                    rut=rut, primer_nombre=primer_nombre, apellido_paterno=apellido_paterno,
                    age=age, sex=sex, episode_id=episode_id, room_location=room_location,
                    clinic_id=clinic_id
                )
                db.session.add(patient)
                db.session.flush()
            else:
                patient.episode_id = episode_id
                patient.room_location = room_location
            
            ticket_id = generate_ticket_id()
            
            ticket = Ticket(
                id=ticket_id, patient_id=patient.id, surgery_id=surgery_id, technique_id=technique_id,
                doctor_id=doctor_id, discharge_slot_id=discharge_slot_id, pavilion_end_time=pavilion_end_time,
                created_by=current_user.username, clinic_id=clinic_id
            )
            
            ticket.set_stay_adjustment_ids(adjustment_ids)
            
            fpa, overnight_stays = ticket.calculate_fpa(pavilion_end_time, technique.base_stay_hours, adjustment_hours, surgery)
            ticket.initial_fpa = fpa
            ticket.current_fpa = fpa
            ticket.overnight_stays = overnight_stays
            
            db.session.add(ticket)
            db.session.commit()
            
            flash(f'Ticket {ticket_id} creado exitosamente.', 'success')
            return redirect(url_for('tickets.detail', ticket_id=ticket_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear el ticket: {str(e)}', 'error')
    
    surgeries = Surgery.query.filter_by(is_active=True, applies_ticket_home=True, clinic_id=clinic_id).all()
    techniques = Technique.query.filter_by(is_active=True, clinic_id=clinic_id).all()
    adjustments = StayAdjustmentCriterion.query.filter_by(is_active=True, clinic_id=clinic_id).all()
    doctors = Doctor.query.filter_by(is_active=True, clinic_id=clinic_id).all()
    discharge_slots = DischargeTimeSlot.query.filter_by(is_active=True, clinic_id=clinic_id).order_by(DischargeTimeSlot.start_time).all()
    
    techniques_json = [t.to_dict() for t in techniques]
    adjustments_json = [a.to_dict() for a in adjustments]
    
    return render_template('tickets/create.html', 
                         surgeries=surgeries, techniques=techniques, techniques_json=techniques_json,
                         adjustments=adjustments, adjustments_json=adjustments_json,
                         doctors=doctors, discharge_slots=discharge_slots)

@tickets_bp.route('/<ticket_id>')
@login_required
def detail(ticket_id):
    ticket = Ticket.query.filter_by(id=ticket_id, clinic_id=current_user.clinic_id).first_or_404()
    
    adjustment_details = []
    if ticket.get_stay_adjustment_ids():
        adjustments = StayAdjustmentCriterion.query.filter(
            StayAdjustmentCriterion.id.in_(ticket.get_stay_adjustment_ids()),
            StayAdjustmentCriterion.clinic_id == current_user.clinic_id
        ).all()
        adjustment_details = adjustments
    
    modification_reasons = StandardizedReason.query.filter_by(category='modification', is_active=True, clinic_id=current_user.clinic_id).all()
    annulment_reasons = StandardizedReason.query.filter_by(category='annulment', is_active=True, clinic_id=current_user.clinic_id).all()
    non_compliance_reasons = StandardizedReason.query.filter_by(category='non_compliance', is_active=True, clinic_id=current_user.clinic_id).all()
    
    return render_template('tickets/detail.html', 
                         ticket=ticket, adjustment_details=adjustment_details,
                         modification_reasons=modification_reasons, annulment_reasons=annulment_reasons,
                         non_compliance_reasons=non_compliance_reasons)

@tickets_bp.route('/<ticket_id>/modify_fpa', methods=['POST'])
@login_required
def modify_fpa(ticket_id):
    ticket = Ticket.query.filter_by(id=ticket_id, clinic_id=current_user.clinic_id).first_or_404()
    
    if not ticket.can_be_modified():
        flash('Este ticket no puede ser modificado.', 'error')
        return redirect(url_for('tickets.detail', ticket_id=ticket_id))
    
    if ticket.get_modification_count() >= 5:
        flash('Se ha alcanzado el límite máximo de modificaciones (5).', 'error')
        return redirect(url_for('tickets.detail', ticket_id=ticket_id))
    
    try:
        new_fpa_str = request.form.get('new_fpa')
        reason = request.form.get('reason')
        justification = request.form.get('justification', '')
        
        if not new_fpa_str or not reason:
            flash('Todos los campos son obligatorios.', 'error')
            return redirect(url_for('tickets.detail', ticket_id=ticket_id))
        
        new_fpa = datetime.strptime(new_fpa_str, '%Y-%m-%dT%H:%M')
        
        modification = FpaModification(
            ticket_id=ticket_id, previous_fpa=ticket.current_fpa, new_fpa=new_fpa,
            reason=reason, justification=justification, modified_by=current_user.username,
            clinic_id=current_user.clinic_id
        )
        
        ticket.current_fpa = new_fpa
        
        time_diff = new_fpa - ticket.pavilion_end_time
        overnight_stays = max(0, time_diff.days)
        if time_diff.seconds > 0:
            overnight_stays += 1
        ticket.overnight_stays = overnight_stays
        
        db.session.add(modification)
        db.session.commit()
        
        flash('FPA modificada exitosamente.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al modificar FPA: {str(e)}', 'error')
    
    return redirect(url_for('tickets.detail', ticket_id=ticket_id))

@tickets_bp.route('/<ticket_id>/close', methods=['POST'])
@login_required
def close_ticket(ticket_id):
    ticket = Ticket.query.filter_by(id=ticket_id, clinic_id=current_user.clinic_id).first_or_404()
    if not ticket.can_be_modified():
        flash('Este ticket no puede ser cerrado.', 'error')
        return redirect(url_for('tickets.detail', ticket_id=ticket_id))
    try:
        actual_discharge_date_str = request.form.get('actual_discharge_date')
        closed_reason = request.form.get('closed_reason')
        
        if not actual_discharge_date_str or not closed_reason:
            flash('Todos los campos son obligatorios para cerrar el ticket.', 'error')
            return redirect(url_for('tickets.detail', ticket_id=ticket_id))
            
        actual_discharge_date = datetime.strptime(actual_discharge_date_str, '%Y-%m-%dT%H:%M')
        
        ticket.status = 'Cerrado'
        ticket.closed_at = datetime.utcnow()
        ticket.actual_discharge_date = actual_discharge_date
        ticket.closed_reason = closed_reason
        
        # Determine compliance status
        if actual_discharge_date <= ticket.current_fpa:
            ticket.compliance_status = 'compliant'
        else:
            ticket.compliance_status = 'non_compliant'
            
        db.session.commit()
        flash(f'Ticket {ticket.id} cerrado exitosamente.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al cerrar el ticket: {str(e)}', 'error')
        
    return redirect(url_for('tickets.detail', ticket_id=ticket_id))

@tickets_bp.route('/<ticket_id>/annul', methods=['POST'])
@login_required
def annul_ticket(ticket_id):
    ticket = Ticket.query.filter_by(id=ticket_id, clinic_id=current_user.clinic_id).first_or_404()
    if not ticket.can_be_modified():
        flash('Este ticket no puede ser anulado.', 'error')
        return redirect(url_for('tickets.detail', ticket_id=ticket_id))
    try:
        annulled_reason = request.form.get('annulled_reason')
        if not annulled_reason:
            flash('La razón de anulación es obligatoria.', 'error')
            return redirect(url_for('tickets.detail', ticket_id=ticket_id))
            
        ticket.status = 'Anulado'
        ticket.annulled_at = datetime.utcnow()
        ticket.annulled_by = current_user.username
        ticket.annulled_reason = annulled_reason
        
        db.session.commit()
        flash(f'Ticket {ticket.id} anulado exitosamente.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al anular el ticket: {str(e)}', 'error')
        
    return redirect(url_for('tickets.detail', ticket_id=ticket_id))


@tickets_bp.route('/')
@login_required
def list():
    filters = {
        'status': request.args.get('status', ''),
        'search': request.args.get('search', ''),
        'surgery': request.args.get('surgery', ''),
        'date_from': request.args.get('date_from', ''),
        'date_to': request.args.get('date_to', ''),
        'compliance': request.args.get('compliance', '')
    }
    sort_by = request.args.get('sort_by', 'created_at')
    sort_dir = request.args.get('sort_dir', 'desc')

    query = _build_tickets_query(filters)

    query = apply_sorting_to_query(query, sort_by, sort_dir)

    tickets = query.all()
    
    for ticket in tickets:
        if ticket.status == 'Vigente' and ticket.current_fpa:
            ticket.time_remaining = calculate_time_remaining(ticket.current_fpa)
        else:
            ticket.time_remaining = None
    
    surgeries = Surgery.query.filter_by(is_active=True, clinic_id=current_user.clinic_id).all()
    
    return render_template('tickets/list.html', 
                         tickets=tickets, 
                         surgeries=surgeries,
                         filters=filters)

@tickets_bp.route('/<ticket_id>/pdf')
@login_required
def export_pdf(ticket_id):
    ticket = Ticket.query.filter_by(id=ticket_id, clinic_id=current_user.clinic_id).first_or_404()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CustomTitle', parent=styles['Title'], fontSize=20, alignment=TA_CENTER, spaceBottom=20))
    styles.add(ParagraphStyle(name='CustomSubTitle', parent=styles['h2'], fontSize=14, alignment=TA_LEFT, spaceBottom=10, spaceTop=10, textColor=colors.HexColor("#14B8A6")))
    styles.add(ParagraphStyle(name='CustomBodyHead', parent=styles['Normal'], alignment=TA_LEFT, textColor=colors.dimgrey, fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle(name='CustomBodyText', parent=styles['Normal'], alignment=TA_LEFT))

    story = []

    # Helper to handle None values
    def p(text, style='CustomBodyText'):
        return Paragraph(str(text or ''), styles[style])

    def h(text, style='CustomBodyHead'):
        return Paragraph(str(text or ''), styles[style])

    # Title
    story.append(p(f"Detalle de Ticket Home: {ticket.id}", 'CustomTitle'))
    story.append(Spacer(1, 0.2*inch))

    # --- Ticket Information ---
    story.append(p("Información del Ticket", 'CustomSubTitle'))
    ticket_data = [
        [h("Estado"), p(ticket.status)],
        [h("FPA Inicial"), p(ticket.initial_fpa.strftime('%d/%m/%Y %H:%M'))],
        [h("FPA Actual"), p(ticket.current_fpa.strftime('%d/%m/%Y %H:%M'))],
        [h("Noches de Estancia"), p(ticket.overnight_stays)],
        [h("Creado por"), p(ticket.created_by)],
        [h("Fecha de Creación"), p(ticket.created_at.strftime('%d/%m/%Y %H:%M'))],
        [h("Bloque Horario de Alta"), p(ticket.discharge_time_slot.name if ticket.discharge_time_slot else 'No asignado')],
    ]
    story.append(Table(ticket_data, colWidths=[2*inch, 4*inch], style=TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
    ])))
    story.append(Spacer(1, 0.2*inch))

    # --- Patient Information ---
    story.append(p("Información del Paciente", 'CustomSubTitle'))
    patient_data = [
        [h("Nombre Completo"), p(ticket.patient.full_name)],
        [h("RUT"), p(ticket.patient.rut)],
        [h("Edad"), p(f"{ticket.patient.age} años")],
        [h("Sexo"), p(ticket.patient.sex)],
    ]
    story.append(Table(patient_data, colWidths=[2*inch, 4*inch], style=TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
    ])))
    story.append(Spacer(1, 0.2*inch))

    # --- Surgery Information ---
    story.append(p("Información Quirúrgica", 'CustomSubTitle'))
    surgery_data = [
        [h("Cirugía"), p(ticket.surgery.name)],
        [h("Técnica"), p(ticket.technique.name)],
        [h("Médico Tratante"), p(ticket.attending_doctor.name if ticket.attending_doctor else 'N/A')],
        [h("Fin de Pabellón"), p(ticket.pavilion_end_time.strftime('%d/%m/%Y %H:%M'))],
    ]
    story.append(Table(surgery_data, colWidths=[2*inch, 4*inch], style=TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
    ])))
    story.append(Spacer(1, 0.2*inch))

    # --- Modification History ---
    if ticket.modifications:
        story.append(p("Historial de Modificaciones", 'CustomSubTitle'))
        for mod in ticket.modifications:
            mod_text = f"""
            <b>Fecha:</b> {mod.modified_at.strftime('%d/%m/%Y %H:%M')} por <b>{mod.modified_by or 'N/A'}</b><br/>
            <b>FPA Anterior:</b> {mod.previous_fpa.strftime('%d/%m/%Y %H:%M')}<br/>
            <b>FPA Nueva:</b> {mod.new_fpa.strftime('%d/%m/%Y %H:%M')}<br/>
            <b>Razón:</b> {mod.reason or 'N/A'}<br/>
            """
            if mod.justification:
                mod_text += f"<b>Justificación:</b> {mod.justification}"
            story.append(p(mod_text))
            story.append(Spacer(1, 0.1*inch))
    
    # --- Closure/Annulment Info ---
    if ticket.status in ['Cerrado', 'Anulado']:
        story.append(p(f"Información de {ticket.status}", 'CustomSubTitle'))
        info_data = []
        if ticket.status == 'Cerrado':
            info_data.extend([
                [h("Fecha de Cierre"), p(ticket.closed_at.strftime('%d/%m/%Y %H:%M') if ticket.closed_at else 'N/A')],
                [h("Fecha Real de Alta"), p(ticket.actual_discharge_date.strftime('%d/%m/%Y %H:%M') if ticket.actual_discharge_date else 'N/A')],
                [h("Cumplimiento"), p(ticket.compliance_status)],
                [h("Razón de Cierre"), p(ticket.closed_reason)],
            ])
        else: # Anulado
            info_data.extend([
                [h("Fecha de Anulación"), p(ticket.annulled_at.strftime('%d/%m/%Y %H:%M') if ticket.annulled_at else 'N/A')],
                [h("Anulado por"), p(ticket.annulled_by)],
                [h("Razón de Anulación"), p(ticket.annulled_reason)],
            ])
        
        story.append(Table(info_data, colWidths=[2*inch, 4*inch], style=TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
        ])))

    doc.build(story)
    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="ticket_{ticket.id}.pdf"'
    return response

@tickets_bp.route('/reports/excel')
@login_required
def export_excel():
    filters = {
        'status': request.args.get('status', ''),
        'surgery': request.args.get('surgery', ''),
        'date_from': request.args.get('date_from', ''),
        'date_to': request.args.get('date_to', ''),
        'compliance': request.args.get('compliance', '')
    }
    query = _build_tickets_query(filters)
    tickets = query.order_by(Ticket.created_at.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Tickets"
    headers = ['N° Ticket', 'Estado', 'RUT Paciente', 'Nombre Completo', 'Cirugía', 'FPA Actual']
    ws.append(headers)
    for ticket in tickets:
        ws.append([ticket.id, ticket.status, ticket.patient.rut, ticket.patient.full_name, ticket.surgery.name, ticket.current_fpa.strftime('%Y-%m-%d %H:%M')])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="reporte_tickets.xlsx"'
    return response


